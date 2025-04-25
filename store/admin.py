from django.contrib import admin
from django.db.models import Sum, F, DecimalField
from django.http import HttpResponse
from django.utils.html import format_html
from django.contrib import messages

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from io import BytesIO

from .models import (
    Product, 
    PurchaseOrder, 
    PurchaseOrderLineItem, 
    Invoice, 
    InvoiceLineItem
)

class StockStatusFilter(admin.SimpleListFilter):
    title = 'Stock Status'
    parameter_name = 'stock_status'
    
    def lookups(self, request, model_admin):
        return (
            ('out', 'Out of Stock'),
            ('low', 'Low Stock (≤ 10)'),
            ('in', 'In Stock (> 10)'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'out':
            return queryset.filter(stock_quantity=0)
        if self.value() == 'low':
            return queryset.filter(stock_quantity__gt=0, stock_quantity__lte=10)
        if self.value() == 'in':
            return queryset.filter(stock_quantity__gt=10)

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('product_image_thumbnail', 'name', 'sku', 'unit_price', 'stock_status', 'inventory_value')
    list_filter = ('created_at', StockStatusFilter)
    search_fields = ('name', 'sku', 'description')
    readonly_fields = ('created_at', 'updated_at', 'product_image_preview', 'inventory_value_display')
    list_per_page = 20
    save_on_top = True
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'sku', 'description', 'image', 'product_image_preview')
        }),
        ('Pricing and Inventory', {
            'fields': ('unit_price', 'stock_quantity', 'inventory_value_display')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    # Custom CSS for admin
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css', 'admin/css/action_fixes.css',)
        }
    
    def product_image_thumbnail(self, obj):
        if obj.image:
            return format_html('<img src="{}" width="50" height="50" style="object-fit: cover; border-radius: 5px;" />', obj.image.url)
        return format_html('<div style="width: 50px; height: 50px; background-color: #f8f9fa; border-radius: 5px; display: flex; align-items: center; justify-content: center;"><i class="fas fa-box" style="color: #adb5bd;"></i></div>')
    product_image_thumbnail.short_description = 'Image'
    
    def product_image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" width="300" style="border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1);" />', obj.image.url)
        return format_html('<div style="width: 300px; height: 200px; background-color: #f8f9fa; border-radius: 8px; display: flex; align-items: center; justify-content: center;"><span style="color: #adb5bd; font-size: 16px;">No image available</span></div>')
    product_image_preview.short_description = 'Image Preview'
    
    def inventory_value(self, obj):
        try:
            value = obj.get_inventory_value()
            if value > 1000:
                return format_html('<span style="color: #28a745; font-weight: bold;">${:,.2f}</span>', value)
            return format_html('${:,.2f}', value)
        except Exception:
            return format_html('$0.00')
    inventory_value.short_description = 'Inventory Value'
    
    def inventory_value_display(self, obj):
        try:
            value = obj.get_inventory_value()
            return format_html('<div style="font-size: 1.5em; color: #28a745; font-weight: bold;">${:,.2f}</div>', value)
        except Exception:
            return format_html('<div style="font-size: 1.5em; color: #6c757d;">$0.00</div>')
    inventory_value_display.short_description = 'Inventory Value'
    
    def stock_status(self, obj):
        if obj.stock_quantity <= 0:
            return format_html('<span class="status-badge out-of-stock">Out of Stock</span>')
        elif obj.stock_quantity <= 10:
            return format_html('<span class="status-badge low-stock">Low Stock ({0})</span>', obj.stock_quantity)
        else:
            return format_html('<span class="status-badge in-stock">In Stock ({0})</span>', obj.stock_quantity)
    stock_status.short_description = 'Stock Status'

    def get_actions(self, request):
        actions = super().get_actions(request)
        actions['export_to_csv'] = (self.export_to_csv, 'export_to_csv', 'Export selected products to CSV')
        return actions

    def export_to_csv(self, modeladmin, request, queryset):
        try:
            import csv
            
            # Set proper content type and filename
            filename = "products.csv"
            content_type = 'text/csv'
            
            # Create the response
            response = HttpResponse(content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Use csv writer for generating the content
            writer = csv.writer(response)
            writer.writerow(['Name', 'SKU', 'Unit Price', 'Stock Quantity', 'Inventory Value'])
            
            # Write the data
            for product in queryset:
                writer.writerow([
                    product.name, 
                    product.sku, 
                    product.unit_price, 
                    product.stock_quantity, 
                    product.get_inventory_value()
                ])
                
            # Add success message
            messages.success(request, f'{queryset.count()} products exported to CSV successfully.')
            return response
            
        except Exception as e:
            messages.error(request, f'Error exporting products to CSV: {str(e)}')
            return None
            
    export_to_csv.short_description = 'Export to CSV'

class PurchaseOrderLineItemInline(admin.TabularInline):
    model = PurchaseOrderLineItem
    min_num = 1
    extra = 0
    fields = ('product', 'quantity', 'cost_per_unit', 'received_quantity', 'subtotal')
    readonly_fields = ('subtotal',)
    
    def subtotal(self, obj):
        if obj.id:
            return f"${obj.get_subtotal()}"
        return "$0.00"
    subtotal.short_description = 'Subtotal'

@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(admin.ModelAdmin):
    list_display = ('order_number', 'vendor_name', 'order_date', 'status', 'total_cost', 'total_items')
    list_filter = ('status', 'order_date')
    search_fields = ('order_number', 'vendor_name', 'notes')
    readonly_fields = ('created_at', 'updated_at', 'total_cost_display')
    inlines = [PurchaseOrderLineItemInline]
    date_hierarchy = 'order_date'
    
    # Add Media class for custom CSS
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css', 'admin/css/action_fixes.css',)
        }
    
    fieldsets = (
        ('Order Information', {
            'fields': ('order_number', 'vendor_name', 'status')
        }),
        ('Dates', {
            'fields': ('order_date', 'expected_delivery_date')
        }),
        ('Additional Information', {
            'fields': ('notes', 'total_cost_display')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def total_cost(self, obj):
        return f"${obj.get_total_cost()}"
    total_cost.short_description = 'Total Cost'
    
    def total_items(self, obj):
        return obj.get_total_items()
    total_items.short_description = 'Total Items'
    
    def total_cost_display(self, obj):
        return format_html('<div style="font-size: 1.2em; color: #28a745; font-weight: bold;">${}</div>', obj.get_total_cost())
    total_cost_display.short_description = 'Total Cost'

class InvoiceLineItemInline(admin.TabularInline):
    model = InvoiceLineItem
    min_num = 1
    extra = 0
    fields = ('product', 'quantity', 'price_each', 'subtotal')
    readonly_fields = ('subtotal',)
    
    def subtotal(self, obj):
        if obj.id:
            return f"${obj.get_subtotal()}"
        return "$0.00"
    subtotal.short_description = 'Subtotal'

@admin.register(Invoice)
class InvoiceAdmin(admin.ModelAdmin):
    list_display = ('invoice_number', 'customer_name', 'invoice_date', 'due_date', 'status_colored', 'total_amount')
    list_filter = ('status', 'invoice_date', 'due_date')
    search_fields = ('invoice_number', 'customer_name', 'customer_email', 'notes')
    readonly_fields = ('created_at', 'updated_at', 'total_amount_display')
    inlines = [InvoiceLineItemInline]
    date_hierarchy = 'invoice_date'
    actions = ['mark_as_paid', 'export_to_xlsx']
    
    # Add Media class for custom CSS
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css', 'admin/css/action_fixes.css',)
        }
    
    fieldsets = (
        ('Invoice Information', {
            'fields': ('invoice_number', 'customer_name', 'customer_email')
        }),
        ('Billing Details', {
            'fields': ('billing_address',)
        }),
        ('Dates and Status', {
            'fields': ('invoice_date', 'due_date', 'status')
        }),
        ('Additional Information', {
            'fields': ('notes', 'total_amount_display')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def total_amount(self, obj):
        return f"${obj.get_total_amount()}"
    total_amount.short_description = 'Total Amount'
    
    def status_colored(self, obj):
        if obj.is_overdue() and obj.status != 'paid' and obj.status != 'cancelled':
            return format_html('<span class="status-badge overdue">OVERDUE</span>')
        
        status_classes = {
            'draft': 'draft',
            'sent': 'sent',
            'paid': 'paid',
            'cancelled': 'cancelled',
            'overdue': 'overdue'
        }
        
        return format_html('<span class="status-badge {}">{}</span>', 
                          status_classes.get(obj.status, 'draft'), 
                          obj.get_status_display())
    status_colored.short_description = 'Status'
    
    def total_amount_display(self, obj):
        return format_html('<div style="font-size: 1.2em; color: #28a745; font-weight: bold;">${}</div>', obj.get_total_amount())
    total_amount_display.short_description = 'Total Amount'
    
    def mark_as_paid(self, request, queryset):
        updated = 0
        for invoice in queryset:
            if invoice.status != 'paid':
                invoice.status = 'paid'
                invoice.save()
                updated += 1
        
        if updated == 0:
            messages.warning(request, 'No invoices were marked as paid.')
        else:
            messages.success(request, f'{updated} {"invoice was" if updated == 1 else "invoices were"} successfully marked as paid.')
    mark_as_paid.short_description = 'Mark selected invoices as paid'
    
    def export_to_xlsx(self, request, queryset):
        """Export invoice details to XLSX file"""
        if queryset.count() != 1:
            messages.warning(request, 'Please select only one invoice to export.')
            return
        
        invoice = queryset.first()
        
        try:
            # Create workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Invoice {invoice.invoice_number}"
            
            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Company header
            ws.merge_cells('A1:F1')
            ws['A1'] = 'E-COMMERCE MANAGEMENT SYSTEM'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Invoice information
            ws['A3'] = 'Invoice Number:'
            ws['B3'] = str(invoice.invoice_number)
            ws['A3'].font = header_font
            
            ws['A4'] = 'Date:'
            ws['B4'] = invoice.invoice_date.strftime('%Y-%m-%d')
            ws['A4'].font = header_font
            
            ws['A5'] = 'Due Date:'
            ws['B5'] = invoice.due_date.strftime('%Y-%m-%d')
            ws['A5'].font = header_font
            
            ws['A6'] = 'Status:'
            ws['B6'] = invoice.get_status_display()
            ws['A6'].font = header_font
            
            # Customer information
            ws['D3'] = 'Customer:'
            ws['E3'] = invoice.customer_name
            ws['D3'].font = header_font
            
            ws['D4'] = 'Email:'
            ws['E4'] = invoice.customer_email
            ws['D4'].font = header_font
            
            ws['D5'] = 'Billing Address:'
            ws['E5'] = invoice.billing_address.replace("\n", ", ")
            ws['D5'].font = header_font
            
            # Line items header
            row = 8
            headers = ['Product', 'SKU', 'Quantity', 'Price Each', 'Subtotal']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Line items
            row += 1
            for item in invoice.line_items.all():
                ws.cell(row=row, column=1, value=item.product.name)
                ws.cell(row=row, column=2, value=item.product.sku)
                ws.cell(row=row, column=3, value=item.quantity)
                ws.cell(row=row, column=4, value=float(item.price_each))
                ws.cell(row=row, column=5, value=float(item.get_subtotal()))
                row += 1
            
            # Total
            ws.cell(row=row+1, column=4, value='Total:')
            ws.cell(row=row+1, column=4).font = header_font
            ws.cell(row=row+1, column=5, value=float(invoice.get_total_amount()))
            ws.cell(row=row+1, column=5).font = header_font
            
            # Adjust column widths
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws.column_dimensions[col].width = 20
            
            # Create response
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Set proper content type and filename
            filename = f"Invoice_{invoice.invoice_number}.xlsx"
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            # Create the HttpResponse object with proper content-type and headers
            response = HttpResponse(
                buffer.getvalue(),
                content_type=content_type
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Add a success message
            messages.success(request, f'Invoice {invoice.invoice_number} has been exported successfully.')
            
            return response
            
        except Exception as e:
            messages.error(request, f'Error exporting invoice: {str(e)}')
            return None
            
    export_to_xlsx.short_description = 'Export to XLSX'
    
# We don't register these models directly as they are used in inlines
# admin.site.register(PurchaseOrderLineItem)
# admin.site.register(InvoiceLineItem)
